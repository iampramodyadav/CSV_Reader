"""
Main Table Editor Class - Part 1
Initialization, UI Setup, Sheet Management
"""
import os
import re
from datetime import datetime
from typing import Dict, Optional
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog
import tksheet
import ttkbootstrap as tb
from formula_engine import FormulaEngine
from ui_components import UIComponents
from custom_helpers import plot_interactive_component_graph_coordinate, get_mdt_dropdown_config, get_file_list_for_column, plot_interactive_component_graph, trim_whitespace, read_label_format, write_label_format,fill_nulls
from config import PLUGIN_FILE_PATH, APP_TITLE, DEFAULT_THEME,DEFAULT_DARK_THEME, LOG_DIR, LOG_FILE, APP_VERSION, FORMULA_HELP_TEXT,SHEET_LIGHT_THEME, SHEET_DARK_THEME, RECENT_FILE
from plugin_manager import load_plugins, run_plugin
import json
from pathlib import Path
import tempfile

from mixins.cf_mixin import CFMixin
from mixins.undo_mixin import UndoMixin
from mixins.formula_mixin import FormulaMixin
from mixins.diff_mixin import DiffMixin
from mixins.split_view_mixin import SplitViewMixin
from mixins.prefs_mixin   import PrefsMixin
from mixins.filter_mixin  import FilterMixin
from mixins.palette_mixin import PaletteMixin
from mixins.extras_mixin import ExtrasMixin

class TableEditor(UndoMixin, CFMixin, FormulaMixin, DiffMixin, SplitViewMixin, PrefsMixin, FilterMixin, PaletteMixin, ExtrasMixin):
    """Main application class for spreadsheet editing."""

    def __init__(self, root, initial_file=None):
        self.root = root
        self.root.title(f"{APP_TITLE} - Untitled")
        
        # Core data
        self.df = pd.DataFrame()
        self.current_file = None
        self.modified = False
        self.dark_mode = False
        self.current_sep = None
        self.txt_sepa = ','
        
        # Multi-sheet support
        self.workbook_sheets: Dict[str, pd.DataFrame] = {}
        self.current_sheet_name = None
        # self.formula_engine = None
        self.formula_cells = {}  # {sheet_name: {row,col: formula}}
        self._init_undo_state()
        
        # ── NEW: per-tab metadata ─────────────────────────────────────────────
        # key  : sheet_name (tab label string)  ← unique because we enforce it
        # value: dict with keys:
        #   file_path   – str | None
        #   sep         – separator string used to load this tab
        #   modified    – bool, dirty flag for THIS tab only
        #   excel_name  – str | None  (base filename if from excel, else None)
        #   is_excel    – bool
        self.tab_meta: Dict[str, dict] = {}
        # ─────────────────────────────────────────────────────────────────────     
        
        # Theme
        self.style = tb.Style(DEFAULT_THEME)
        
        # Build UI
        self._setup_ui()
        self._setup_plugin_menu()
        self._load_prefs()                    # ← PrefsMixin: loads and applies saved settings
        self._register_palette_commands()     # ← PaletteMixin: builds the command list
        self._setup_extras()               # ← ADD
        # Load initial file or show empty sheet
        if initial_file and os.path.exists(initial_file):
            self.load_file_guarded(initial_file)
        else:
            try:
                prefs = getattr(self, '_prefs')
                startup_file = prefs.get("startup_file", "")
                if startup_file.strip() != "":
                    self.load_file_guarded(startup_file)
                else:
                    self.update_sheet_from_dataframe()
            except:
                self.update_sheet_from_dataframe()

    def _setup_ui(self):
        """Setup all UI components."""
        UIComponents.create_menu_bar(self.root, self)
        UIComponents.create_toolbar(self.root, self)
        self._build_filter_bar(self.root)
        # Content frame with notebook
        # content_frame = tb.Frame(self.root)
        # content_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        # self.sheet_notebook = ttk.Notebook(content_frame)
        # self.sheet_notebook.pack(fill=tk.BOTH, expand=True)
        content_frame = tb.Frame(self.root)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self._split_content_frame = content_frame          # store ref for mixin
        self.sheet_notebook = self._build_left_pane(content_frame)  # mixin builds it
                
        self.sheet_notebook.bind('<<NotebookTabChanged>>', self._on_sheet_change)
        
        # Create first sheet
        self._create_sheet_tab("Sheet1")
        
        UIComponents.create_formula_bar(self.root, self)
        UIComponents.create_status_bar(self.root, self)
    def _setup_plugin_menu(self):
        """Build the Plugins menu from the plugin file."""
        # import tkinter as tk

        # Find the existing menu bar
        menubar = self.root["menu"]
        if not menubar:
            return
        menubar_widget = self.root.nametowidget(menubar)

        # Create Plugins top-level menu
        self._plugin_menu = tk.Menu(menubar_widget, tearoff=0)
        menubar_widget.add_cascade(label="Plugins", menu=self._plugin_menu)

        # Add reload + open-file utilities
        self._plugin_menu.add_command(
            label="⟳  Reload Plugins",
            command=self._reload_plugin_menu
        )
        self._plugin_menu.add_command(
            label="📂  Open Plugin File",
            command=lambda: self._open_plugin_file(PLUGIN_FILE_PATH)
        )
        self._plugin_menu.add_separator()

        # Load and populate plugin entries
        self._populate_plugin_entries()


    def _populate_plugin_entries(self):
        """Clear and re-add all plugin entries from the plugin file."""

        # Remove everything after the first 3 fixed items (Reload, Open, separator)
        end = self._plugin_menu.index("end")
        if end is not None and end >= 3:
            for i in range(end, 2, -1):  # remove from end down to index 3
                self._plugin_menu.delete(i)

        plugins = load_plugins()
        if not plugins:
            self._plugin_menu.add_command(
                label="(No plugins found — check plugin file path)",
                state="disabled"
            )
            return

        for label in plugins:
            self._plugin_menu.add_command(
                label=label,
                command=lambda lbl=label: self._run_plugin(lbl)
            )


    def _reload_plugin_menu(self):
        """Reload plugin file and rebuild menu."""
        self._populate_plugin_entries()
        self.set_status("Plugins reloaded.")


    def _open_plugin_file(self, path):
        """Open the plugin file in the default editor (Notepad, VS Code, etc.)."""
        import subprocess, os
        if os.path.exists(path):
            # os.startfile(path)   # Windows; use subprocess.run(["xdg-open", path]) on Linux
            npp_path = r"C:\Program Files\Notepad++\notepad++.exe"

            if os.path.exists(npp_path):
                subprocess.Popen([npp_path, path])
            else:
                print("Notepad++ not found. Falling back to system Notepad.")
                subprocess.Popen(['notepad.exe', path])
        else:
            messagebox.showwarning(
                "Plugin file not found",
                f"Plugin file does not exist:\n{path}\n\nCreate it to get started."
            )


    def _run_plugin(self, label):
        """Sync sheet → df, run plugin, then push result back to sheet."""
        # from plugin_manager import run_plugin
        self._push_undo() 
        # Always sync latest sheet data into df first
        self.update_dataframe_from_sheet()
        
        df_result = run_plugin(label, self.df, self)
        
        # If plugin returned a modified df, push it back to the sheet
        if df_result is not None and not df_result.equals(self.df):
            self.df = df_result
            self.workbook_sheets[self.current_sheet_name] = self.df
            self.update_sheet_from_dataframe()
            self.mark_modified()
            
    def _create_sheet_tab(self, sheet_name: str, df: Optional[pd.DataFrame] = None,
                          file_path: str = None, sep: str = ',',
                          excel_name: str = None, is_excel: bool = False):
        """Create a new sheet tab. Now stores per-tab file metadata."""
        
        # ── Make sheet_name unique across ALL currently open tabs ─────────────
        base = sheet_name
        counter = 1
        while sheet_name in self.workbook_sheets:
            sheet_name = f"{base}_{counter}"
            counter += 1
    
        tab_frame = tb.Frame(self.sheet_notebook)
        self.sheet_notebook.add(tab_frame, text=sheet_name)
        
        sheet = tksheet.Sheet(
            tab_frame, show_x_scrollbar=True, show_y_scrollbar=True,
            show_top_left=True, headers=None
        )
        sheet.enable_bindings(("single_select", "row_select", "column_select", "drag_select",
                              "edit_cell", "edit_header", "copy", "cut", "paste", "delete",
                              "undo", "redo", "find", "replace", "replace_all","arrowkeys",
                              "sort_columns", "sort_rows", "rc_insert_row", "rc_delete_row",
                              "row_height_resize", "column_height_resize", "column_width_resize",
                              "row_width_resize", "rc_insert_column", "rc_delete_column", "double_click_row_resize", 
                              "ctrl_select", "double_click_column_resize","move_columns","move_rows", "arrowkeys", "select_all"))
        
        sheet.disable_bindings("undo", "redo")          # removes Ctrl+Z/Y from tksheet
        self._bind_sheet_modified(sheet, sheet_name)    # wire our handler
        
        sheet.pack(fill=tk.BOTH, expand=True)
        
        sheet.extra_bindings([
            ("rc_insert_column", self.add_column),
            ("rc_insert_row", self.insert_row),
            ("cell_select", self.update_selection_status),
            ("drag_select_cells", self.update_selection_status),
            ("end_edit_cell", self.on_cell_edit),
            # ("end_edit_cell",    self._filter_sync_edit),   # ← ADD THIS
            ("end_edit_header", self.mark_modified),
            ("rc_delete_column", self.mark_modified),
            ("rc_delete_row", self.mark_modified),
            ("paste", self.mark_modified),
            ("delete", self.mark_modified),
            ("cut", self.mark_modified),
            ("replace", self.mark_modified),
            ("replace_all", self.mark_modified),
            ("sort_columns", self.mark_modified),
            ("move_columns", self.mark_modified),
            ("move_rows", self.mark_modified),
            # ("undo", self.mark_modified),
            # ("redo", self.mark_modified),
        ])
        
        sheet.bind("<Control-r>", self.autofill_selection)
        self.bind_navigation_shortcuts(sheet)
        
        sheet.popup_menu_add_command("Extract Rows to New Tab",self.extract_selection_to_new_tab)
        sheet.popup_menu_add_command("Open Containing Folder",self.open_containing_folder)
        sheet.popup_menu_add_command("Copy Full Path",        self.copy_full_path)
        sheet.popup_menu_add_command("Copy File Name",        self.copy_file_name)      
        sheet.popup_menu_add_command("Rename Sheet", self.rename_sheet)
        sheet.popup_menu_add_command("Close This Tab", self.close_current_tab)
        sheet.popup_menu_add_command('Column Stats', self.show_column_stats)
        sheet.popup_menu_add_command("Autofill (Ctrl+R)", self.autofill_selection)

        
        # Store DataFrame
        # print(df)
        self.workbook_sheets[sheet_name] = df if df is not None else pd.DataFrame()
        # print(df)
        self.current_sheet_name = sheet_name
        
        # ── NEW: store per-tab metadata ───────────────────────────────────────
        self.tab_meta[sheet_name] = {
            "file_path": file_path,
            "sep":       sep,
            "modified":  False,
            "excel_name": excel_name,
            "is_excel":   is_excel,
        }
        # ─────────────────────────────────────────────────────────────────────
        
        if not hasattr(self, 'formula_engine') or self.formula_engine is None:
            self.formula_engine = FormulaEngine(sheet)
        
        return sheet

    def get_current_sheet(self) -> tksheet.Sheet:
        """Get currently active sheet."""
        current_tab = self.sheet_notebook.select()
        if not current_tab:
            return None
        tab_frame = self.sheet_notebook.nametowidget(current_tab)
        for widget in tab_frame.winfo_children():
            if isinstance(widget, tksheet.Sheet):
                return widget
        return None

    def rename_sheet(self):
            """Opens a dialog to rename the currently selected sheet (notebook tab)."""
            current_sheet_name = self.current_sheet_name
            if not current_sheet_name:
                messagebox.showinfo("Rename Sheet", "No sheet is currently selected.")
                return
            
            # 1. Ask for new name
            new_name = simpledialog.askstring(
                "Rename Sheet", 
                f"Enter new name for '{current_sheet_name}':", 
                initialvalue=current_sheet_name
            )

            if not new_name or new_name.strip() == current_sheet_name:
                self.set_status(f"Entered the same sheet name {self.current_sheet_name}")
                return # User canceled or entered the same name
            
            new_name = new_name.strip()

            # 2. Check for uniqueness and validity
            if new_name in self.workbook_sheets:
                messagebox.showerror("Error", f"Sheet name '{new_name}' already exists.")
                return
            
            if not new_name:
                messagebox.showerror("Error", "Sheet name cannot be empty.")
                return

            # 3. Update the Notebook tab
            try:
                # current_tab_id = self.notebook.select()
                current_tab_id = self.sheet_notebook.select()
                self.sheet_notebook.tab(current_tab_id, text=new_name)
                
                # 4. Update the internal dictionary and references
                sheet_data = self.workbook_sheets.pop(current_sheet_name) # Remove with old key
                sheet_data.name = new_name                       # Update name in SheetData object
                self.workbook_sheets[new_name] = sheet_data               # Insert with new key
                # Keep tab_meta in sync with renamed sheet
                if current_sheet_name in self.tab_meta:
                    self.tab_meta[new_name] = self.tab_meta.pop(current_sheet_name)
                if current_sheet_name in self.formula_cells:
                    self.formula_cells[new_name] = self.formula_cells.pop(current_sheet_name)
                self.current_sheet_name = new_name               # Update the current reference
                
                self.set_status(f"Sheet renamed from '{current_sheet_name}' to '{new_name}'.")

            except Exception as e:
                messagebox.showerror("Error", f"Failed to rename sheet: {e}")

    def _on_sheet_change(self, event=None):
        """Handle sheet tab change — sync all per-tab globals."""
        sheet = self.get_current_sheet()
        if sheet:
            self.current_sheet_name = self.sheet_notebook.tab(self.sheet_notebook.select(), "text")
            # print("_on_sheet_change-_sync_globals_from_current_tab")
            self._sync_globals_from_current_tab()          # ← replaces the old manual df/formula lines
            # Guard: formula_engine may not exist yet during startup
            if self.formula_engine is not None:
                self.formula_engine.sheet = sheet
            self.update_selection_status()
            # Show which file this tab belongs to in status
            meta = self.tab_meta.get(self.current_sheet_name, {})
            fp = meta.get("file_path")
            label = os.path.basename(fp) if fp else "Unsaved"
            self.set_status(f"Tab: {self.current_sheet_name}  |  File: {label}")
            if hasattr(self, 'apply_cf_rules'):
                self.apply_cf_rules()
            self.update_title()
            self._refresh_right_pane()
            self._on_filter_tab_change() 
    
    def _save_sheet_as_separate_file(self, sheet_name: str):
        """Save new sheet as separate file."""
        current_dir = os.path.dirname(self.current_file) if self.current_file else os.getcwd()
        
        file_path = filedialog.asksaveasfilename(
            initialdir=current_dir,
            initialfile=f"{sheet_name}.csv",
            defaultextension=".csv",
            filetypes=[
                ("CSV", "*.csv"),
                ("TSV", "*.tsv"),
                ("Text", "*.txt"),
                ("Excel", "*.xlsx")
            ],
            title=f"Save '{sheet_name}' as separate file"
        )
        
        if not file_path:
            return
        
        try:
            df = pd.DataFrame("",index=range(5), columns=["A", "B", "C", "D","E"])
            
            if file_path.endswith(('.xlsx', '.xls')):
                df.to_excel(file_path, index=False, sheet_name=sheet_name)
            elif file_path.endswith('.tsv'):
                df.to_csv(file_path, sep='\t', index=False)
            else:
                df.to_csv(file_path, sep=self.txt_sepa or ',', index=False)
            
            messagebox.showinfo("Success", f"New sheet saved as:\n{file_path}")
            
            if messagebox.askyesno("Open File?", "Do you want to open the new file?"):
                self.load_file_guarded(file_path)
        
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save separate file: {e}")

    def _populate_sheet(self, sheet_widget: tksheet.Sheet, df: pd.DataFrame):
        """Populate sheet with DataFrame."""
        if df.empty:
            sheet_widget.headers(["A", "B", "C", "D","E"])
            sheet_widget.set_sheet_data([["", "", "", "", ""], ["", "", "", "", ""]])
        else:
            # print(f"_populate_sheet {df}")
            sheet_widget.headers([str(c) for c in df.columns])
            sheet_widget.set_sheet_data(df.astype(str).values.tolist())
        
        self.toggle_dark_mode()
        sheet_widget.refresh()

    def update_sheet_from_dataframe(self):
        """Refresh table display."""
        sheet = self.get_current_sheet()
        if not sheet:
            return
        
        if self.df.empty:
            sheet.headers(["A", "B", "C", "D","E"])
            sheet.set_sheet_data([  ["", "", "", "", ""],
                                    ["", "", "", "", ""],
                                    ["", "", "", "", ""], 
                                    ["", "", "", "", ""], 
                                    ["", "", "", "", ""]])
            self.update_dataframe_from_sheet()

        else:
            sheet.headers([str(c) for c in self.df.columns])
            sheet.set_sheet_data(self.df.astype(str).values.tolist())

        self.workbook_sheets[self.current_sheet_name] = self.df
        
        self.toggle_dark_mode()
        sheet.refresh()
        self.apply_cf_rules()
        self._refresh_pre_edit_snapshot()
        if hasattr(self, 'apply_cf_rules'):
            self.apply_cf_rules()
        self._refresh_right_pane()
        self._on_df_changed_outside_filter()
        
    def update_dataframe_from_sheet(self):
        """Sync DataFrame from sheet."""
        sheet = self.get_current_sheet()
        if not sheet:
            return
        if hasattr(self, '_filter_is_active') and self._filter_is_active():
            return                        # ← ADD THESE TWO LINES
        self.df = pd.DataFrame(sheet.get_sheet_data(), columns=sheet.headers())
        self.workbook_sheets[self.current_sheet_name] = self.df

    def toggle_dark_mode(self):
        """Apply dark/light app theme. Sheet theme is set independently via Preferences."""
        self.dark_mode = self.dark_mode_var.get()
    
        # ── App (ttkbootstrap) theme ──────────────────────────────────────────
        # Use theme from preferences if available, else fall back to config defaults
        if hasattr(self, '_prefs'):
            app_theme = self._prefs.get("theme", DEFAULT_THEME)
            # If the saved theme clashes with dark_mode state, pick sensibly
            from mixins.prefs_mixin import THEMES_DARK
            if self.dark_mode and app_theme not in THEMES_DARK:
                app_theme = DEFAULT_DARK_THEME
            elif not self.dark_mode and app_theme in THEMES_DARK:
                app_theme = DEFAULT_THEME
        else:
            app_theme = DEFAULT_DARK_THEME if self.dark_mode else DEFAULT_THEME
    
        self.style.theme_use(app_theme)
    
        # ── Sheet (tksheet) theme — from preferences, not hardcoded ──────────
        if hasattr(self, '_apply_sheet_theme_from_prefs'):
            self._apply_sheet_theme_from_prefs()
        else:
            # Fallback if prefs mixin not loaded
            sheet_theme = SHEET_DARK_THEME if self.dark_mode else SHEET_LIGHT_THEME
            for tab_id in self.sheet_notebook.tabs():
                tab_frame = self.sheet_notebook.nametowidget(tab_id)
                for widget in tab_frame.winfo_children():
                    if isinstance(widget, tksheet.Sheet):
                        widget.set_options(theme=sheet_theme, align="center")
                        widget.refresh()
    
        # ── Status bar colours ────────────────────────────────────────────────
        status_bg, status_fg = ("#2e2e2e", "white") if self.dark_mode else ("#f0f0f0", "black")
        self.status_frame.config(bg=status_bg)
        self.left_status.config(bg=status_bg, fg=status_fg)
        self.right_status.config(bg=status_bg, fg=status_fg)
        self.select_status.config(bg=status_bg, fg=status_fg)
        self.cell_info.config(bg=status_bg, fg=status_fg)
    
        # ── Sync split view right pane if open ────────────────────────────────
        if hasattr(self, 'update_right_pane_theme'):
            self.update_right_pane_theme()

    def mark_modified(self, event=None):
        """Mark THIS tab as modified (not the whole session)."""
        # Per-tab flag
        if self.current_sheet_name in self.tab_meta:
            self.tab_meta[self.current_sheet_name]["modified"] = True
    
        # Global flag = True if any tab is dirty
        self.modified = any(m["modified"] for m in self.tab_meta.values())
    
        self.update_title()
        # self._push_undo() 
        self.update_dataframe_from_sheet()
        if self.suggestion_mode_var.get():
            self.clean_dropdown_value()
            self.customize_turbine_columns()
    
        sheet = self.get_current_sheet()
        if sheet:
            sheet.refresh()
        self.set_status("Modified")
        
        if hasattr(self, 'apply_cf_rules'):
            self.apply_cf_rules()
        self._refresh_right_pane()
        
    def update_title(self):
        """Title = AppName - active_file [*dirty_tabs*]"""
        meta = self.tab_meta.get(self.current_sheet_name, {})
        fp = meta.get("file_path")
        modif = meta.get("modified")
        # name = os.path.basename(fp) if fp else "Untitled"
        name = fp if fp else "Untitled"
        dirty_count = sum(1 for m in self.tab_meta.values() if m.get("modified"))
        dirty_str = f"  [{dirty_count} unsaved]" if dirty_count else ""
        if modif:
            self.root.title(f"{APP_TITLE} - {name}*{dirty_str}")
        else:
            self.root.title(f"{APP_TITLE} - {name}{dirty_str}")
            
    def set_status(self, text):
        """Update status bar."""
        self.right_status.config(text=text)

    def set_sel_status(self, text):
        """Update selection status."""
        self.select_status.config(text=text)

    def log_usage(self, file_path):
        """Log usage."""
        try:
            log_file = os.path.join(LOG_DIR, LOG_FILE)
            with open(log_file, "a", encoding="utf-8") as f:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                user = os.getlogin()
                f.write(f"{timestamp} | {user} | {file_path} | {APP_VERSION}\n")
        except:
            pass

    def ask_save_changes(self, action="continue"):
        """Ask to save changes."""
        result = messagebox.askyesnocancel("Unsaved Changes", 
                                          f"Save changes before {action}?")
        if result is True:
            self.save_file()
            return True
        if result is None:
            return False
        return True

    def on_close(self):
        """Check all tabs for unsaved changes before exit."""
        dirty_tabs = [name for name, m in self.tab_meta.items() if m.get("modified")]
        if dirty_tabs:
            names = "\n".join(f"  • {t}" for t in dirty_tabs)
            result = messagebox.askyesnocancel(
                "Unsaved Changes",
                f"These tabs have unsaved changes:\n{names}\n\nSave before exiting?"
            )
            if result is None:    # Cancel
                return
            if result is True:    # Save all
                # Save each unique file
                saved = set()
                original_tab = self.current_sheet_name
                for tab_name in dirty_tabs:
                    fp = self.tab_meta[tab_name].get("file_path")
                    if fp and fp not in saved:
                        # Temporarily switch context to save correctly
                        self.current_sheet_name = tab_name
                        self._sync_globals_from_current_tab()
                        self.save_file()
                        saved.add(fp)
                self.current_sheet_name = original_tab
        self.root.destroy()

    def load_file(self, file_path: str, force_sep: str = None):
        """
        Load a file and ADD its sheets to the session.
        Never clears existing tabs — each call appends new tabs.
        force_sep: override auto-detection (used by "Add File" with separator choice).
        """
        try:
            sep_to_use = force_sep #or self.current_sep
            # ── NEW: remove the empty startup Sheet1 if it's the only tab
            # and has never been modified (no file, empty df).
            # This prevents it from polluting the undo history.
            self._remove_empty_startup_sheet()   # ← ADD THIS CALL
            
            if file_path.endswith(('.xlsx', '.xls')):
                excel_file = pd.ExcelFile(file_path)
                excel_base = os.path.basename(file_path)          # e.g. "data.xlsx"
                for sheet_name in excel_file.sheet_names:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name, keep_default_na=False)
                    # Tab label = "data.xlsx » Sheet1"  — clearly shows which file it came from
                    tab_label = f"{os.path.splitext(excel_base)[0]} » {sheet_name}"
                    sheet_widget = self._create_sheet_tab(
                        tab_label, df,
                        file_path=file_path, sep=None,
                        excel_name=excel_base, is_excel=True
                    )
                    self._populate_sheet(sheet_widget, df)
                    # ── Seed snapshot AFTER populate so it has real data ──────
                    self._refresh_pre_edit_snapshot(tab_label)   # ← ADD
                # Switch to first newly added Excel sheet
                # Find last tab index and jump to first of newly added batch
                self.sheet_notebook.select(len(self.sheet_notebook.tabs()) - len(excel_file.sheet_names))
    
            else:
                # ── Auto-detect separator ─────────────────────────────────────
                with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                    first_line = f.readline()
    
                if sep_to_use == "#label" or (not sep_to_use and "#label" in first_line.lower()):
                    _, _, _, df = read_label_format(file_path)
                    detected_sep = "#label"
                elif sep_to_use:
                    df = pd.read_csv(file_path, sep=sep_to_use, engine="python",
                                     keep_default_na=False, na_values=[])
                    detected_sep = sep_to_use
                elif file_path.endswith(".tsv") or "\t" in first_line:
                    df = pd.read_csv(file_path, sep="\t", engine="python",
                                     keep_default_na=False, na_values=[])
                    detected_sep = "\t"
                elif ";" in first_line:
                    df = pd.read_csv(file_path, sep=";", engine="python",
                                     keep_default_na=False, na_values=[])
                    detected_sep = ";"
                elif "," in first_line:
                    df = pd.read_csv(file_path, sep=",", engine="python",
                                     keep_default_na=False, na_values=[])
                    detected_sep = ","
                elif file_path.endswith(".txt") and " " in first_line:
                    df = pd.read_csv(file_path, sep=r"\s+", engine="python",
                                     keep_default_na=False, na_values=[])
                    detected_sep = r"\s+"
                else:
                    df = pd.read_csv(file_path, keep_default_na=False, na_values=[])
                    detected_sep = ","
    
                sheet_name = os.path.splitext(os.path.basename(file_path))[0]
                sheet_widget = self._create_sheet_tab(
                    sheet_name, df,
                    file_path=file_path, sep=detected_sep,
                    excel_name=None, is_excel=False
                )
                self._populate_sheet(sheet_widget, df)
                # ── Seed snapshot AFTER populate ──────────────────────────────
                self._refresh_pre_edit_snapshot(sheet_name)   # ← ADD
                # Switch to the newly added tab
                self.sheet_notebook.select(len(self.sheet_notebook.tabs()) - 1)
    
            # Update globals to reflect current tab
            self._sync_globals_from_current_tab()
            self.update_title()
            if self.suggestion_mode_var.get():
                self.customize_turbine_columns()
            self.set_status(f"Opened: {os.path.basename(file_path)}")
            self._sv_update_combo()
            self.log_usage(file_path)
    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open file: {e}")
            self.set_status("Error opening file")
            
    def _sync_globals_from_current_tab(self):
        """
        Update self.current_file and self.txt_sepa from the active tab's metadata.
        Call this whenever the active tab changes.
        """
        meta = self.tab_meta.get(self.current_sheet_name)
        if meta:
            self.current_file = meta["file_path"]
            self.txt_sepa     = meta["sep"] or ','
        else:
            self.current_file = None
            self.txt_sepa     = ','
        # Also update self.df
        self.df = self.workbook_sheets.get(self.current_sheet_name, pd.DataFrame())
        # print(f"_sync_globals_from_current_tab {self.current_sheet_name} {self.df}")
        
    def save_file(self):
        """
        Save the currently active tab.
        - Non-Excel tab  → saves just that file.
        - Excel tab      → saves ALL tabs that belong to the same .xlsx file together.
        """
        meta = self.tab_meta.get(self.current_sheet_name)
        if not meta or not meta.get("file_path"):
            self.save_file_as()
            return
    
        self.update_dataframe_from_sheet()
        file_path = meta["file_path"]
    
        try:
            if meta.get("is_excel"):
                # Collect ALL tabs that share this exact excel file_path
                excel_sheets = {
                    sname: self.workbook_sheets[sname]
                    for sname, m in self.tab_meta.items()
                    if m.get("file_path") == file_path and sname in self.workbook_sheets
                }
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    for tab_label, df in excel_sheets.items():
                        # Strip the "filename » " prefix to get the real sheet name
                        real_sheet = tab_label.split(" » ", 1)[-1] if " » " in tab_label else tab_label
                        df_save = df.copy()
                        df_save.replace(r'^\s*', '', regex=True, inplace=True)
                        df_save.to_excel(writer, sheet_name=real_sheet, index=False)
                # Mark all those tabs as clean
                for sname, m in self.tab_meta.items():
                    if m.get("file_path") == file_path:
                        m["modified"] = False
            else:
                df_to_save = self.df.copy()
                sep = meta.get("sep", ",")
                df_to_save.replace(r'^\s*$', 'NaN', regex=True, inplace=True)
                if sep == "#label":
                    write_label_format(file_path, self.df, sep="::")
                elif sep == r'\s+' or (sep and re.match(r'^\s+$', sep)):
                    df_to_save.to_string(buf=file_path, index=False, col_space=15, justify='left')
                elif file_path.endswith(".tsv"):
                    df_to_save.to_csv(file_path, sep="\t", index=False)
                else:
                    self.df.to_csv(file_path, sep=sep, index=False, na_rep="")
                meta["modified"] = False
    
            # Recompute global modified
            self.modified = any(m["modified"] for m in self.tab_meta.values())
            self.update_title()
            self.set_status(f"Saved: {os.path.basename(file_path)}")
            # messagebox.showinfo("Saved", "File saved successfully!")
    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save: {e}")
            self.set_status("Error saving file")

    def save_file_as(self):
        """Save current tab under a new filename."""
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[
                ("Excel Workbook", "*.xlsx"),
                ("CSV", "*.csv"),
                ("TSV", "*.tsv"),
                ("Text", "*.txt")
            ]
        )
        if not file_path:
            return
        # Update this tab's metadata to the new path
        if self.current_sheet_name in self.tab_meta:
            self.tab_meta[self.current_sheet_name]["file_path"] = file_path
            self.tab_meta[self.current_sheet_name]["is_excel"] = file_path.endswith(('.xlsx', '.xls'))
            if not file_path.endswith(('.xlsx', '.xls')):
                # Infer sep from extension
                if file_path.endswith('.tsv'):
                    self.tab_meta[self.current_sheet_name]["sep"] = "\t"
                else:
                    self.tab_meta[self.current_sheet_name]["sep"] = ","
        self.current_file = file_path
        self.save_file()

    def export_to_excel(self):
        """Export to Excel with formatting."""
        try:
            #import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("Error", "openpyxl not installed. Using basic export.")
            self.save_file_as()
            return
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")]
        )
        if not file_path:
            return
        
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, df in self.workbook_sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    worksheet = writer.sheets[sheet_name]
                    
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
            
            self.set_status(f"Exported to: {os.path.basename(file_path)}")
            messagebox.showinfo("Success", "Excel file exported with formatting!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {e}")
            
    def export_to_csv(self):
        """Export to Excel with formatting."""
        
        file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")])
        if not file_path:
            return
        
        try:
            self.df.to_csv(file_path, sep=',', index=False)
            self.set_status(f"Exported to: {os.path.basename(file_path)}")
            messagebox.showinfo("Success", "CSV file exported with formatting!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {e}")
            
    def export_to_tsv(self):
        """Export to Excel with formatting."""
        
        file_path = filedialog.asksaveasfilename(defaultextension=".tsv", filetypes=[("TSV", "*.tsv")])
        if not file_path:
            return
        
        try:
            self.df.to_csv(file_path, sep='\t', index=False)
            self.set_status(f"Exported to: {os.path.basename(file_path)}")
            messagebox.showinfo("Success", "TSV file exported with formatting!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {e}")
            
    def export_to_txt(self):
        """Export to Excel with formatting."""
        
        file_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text", "*.txt *.dat")])
        if not file_path:
            return
            
        sep = simpledialog.askstring("Export Options", "Enter separator (use \\t for tab):", 
                                       initialvalue=",")
        if not sep:
            return
            
        if sep == "\\t": sep = "\t"
        
        try:
            self.df.to_csv(file_path, sep=sep, index=False)
            self.set_status(f"Exported to: {os.path.basename(file_path)}")
            messagebox.showinfo("Success", "Text file exported with formatting!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {e}")
            
    def export_to_lib(self):
        """Export to Excel with formatting."""
        
        file_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Library Label", "*.txt")])
        if not file_path:
            return
        try:
            write_label_format(file_path, self.df, sep="::")
            self.set_status(f"Exported to: {os.path.basename(file_path)}")
            messagebox.showinfo("Success", "Library label file exported with formatting!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {e}") 
            
    def export_to_md(self):
        """Export to Excel with formatting."""
        
        file_path = filedialog.asksaveasfilename(defaultextension=".md", filetypes=[("Markdown", "*.md")])
        if not file_path:
            return
        
        try:
            self.df.to_markdown(file_path, index=False)
            self.set_status(f"Exported to: {os.path.basename(file_path)}")
            messagebox.showinfo("Success", "Markdown file exported with formatting!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {e}")
            
    def export_to_html(self):
        """Export to Excel with formatting."""
        
        file_path = filedialog.asksaveasfilename(defaultextension=".html", filetypes=[("HTML", "*.html")])
        if not file_path:
            return
        
        try:
            self.df.to_html(file_path, index=False)
            self.set_status(f"Exported to: {os.path.basename(file_path)}")
            messagebox.showinfo("Success", "HTML file exported with formatting!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {e}")
            
    def export_to_json(self):
        """Export to Excel with formatting."""
        
        file_path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON", "*.json")])
        if not file_path:
            return
        
        try:
            self.df.to_json(file_path, orient='records', indent=4)
            self.set_status(f"Exported to: {os.path.basename(file_path)}")
            messagebox.showinfo("Success", "JSON file exported with formatting!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {e}")
            
    def export_to_xml(self):
        """Export to Excel with formatting."""
        
        file_path = filedialog.asksaveasfilename(defaultextension=".xml", filetypes=[("XML", "*.xml")])
        if not file_path:
            return
        
        try:
            self.df.to_xml(file_path, index=False)
            self.set_status(f"Exported to: {os.path.basename(file_path)}")
            messagebox.showinfo("Success", "XML file exported with formatting!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {e}")
            
    def export_to_tex(self):
        """Export to Excel with formatting."""
        
        file_path = filedialog.asksaveasfilename(defaultextension=".tex", filetypes=[("LaTeX", "*.tex")])
        if not file_path:
            return
        
        try:
            self.df.to_latex(file_path, index=False)
            self.set_status(f"Exported to: {os.path.basename(file_path)}")
            messagebox.showinfo("Success", "LaTeX file exported with formatting!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {e}")
            
    def open_file(self):
        """
        Open file dialog. If tabs already exist, ask:
          - Add to session (append tabs)
          - Replace session (clear and open fresh)
        """
        file_path = filedialog.askopenfilename(
            filetypes=[
                ("All Supported", "*.csv *.txt *.tsv *.xlsx *.xls"),
                ("Excel files",   "*.xlsx *.xls"),
                ("CSV/Text",      "*.csv *.txt *.tsv"),
                ("All files",     "*.*"),
            ]
        )
        if not file_path:
            return
    
        has_existing = bool(self.tab_meta)   # any tabs already open?
    
        if has_existing:
            choice = messagebox.askyesnocancel(
                "Open File",
                f"Add  '{os.path.basename(file_path)}'  to current session?\n\n"
                "YES\t= Add as new tab(s)  ← keeps everything open\n"
                "NO\t= Close all tabs and open fresh\n"
                "CANCEL\t= Abort"
            )
            if choice is None:      # Cancel
                return
            if choice is False:     # No = replace session
                if self.modified:
                    if not self.ask_save_changes("close the current session"):
                        return
                self._close_all_tabs()
    
        self.load_file_guarded(file_path)
        self._save_recent(file_path)

    
    def add_file(self):
        """
        Dedicated 'Add File' button/menu item.
        Always appends — never asks, never clears.
        Lets user choose separator before loading.
        """
        file_path = filedialog.askopenfilename(
            title="Add File to Session",
            filetypes=[
                ("All Supported", "*.csv *.txt *.tsv *.xlsx *.xls"),
                ("All files",     "*.*"),
            ]
        )
        if not file_path:
            return

        # For non-Excel, let user confirm or override separator
        if not file_path.endswith(('.xlsx', '.xls')):
            # sep_choice = simpledialog.askstring(
            #     "Separator",
            #     f"Separator for  {os.path.basename(file_path)}\n"
            #     "Leave blank = auto-detect\n"
            #     "Options: ,   ;   \\t   |   \\s+   #label",
            #     initialvalue=""
            # )
            force_sep = None
            # if sep_choice and sep_choice.strip():
            #     s = sep_choice.strip()
            #     if s.lower() in ["\\t", "tab"]:
            #         force_sep = "\t"
            #     elif s.lower() in ["none", "auto", ""]:
            #         force_sep = None
            #     else:
            #         force_sep = s
            self.load_file_guarded(file_path, force_sep=force_sep)
            self._save_recent(file_path)
        else:
            self.load_file_guarded(file_path)
            self._save_recent(file_path)

    def _close_all_tabs(self):
        """Clear all tabs, data, and metadata. Used when replacing the session."""
        for tab_id in self.sheet_notebook.tabs():
            self.sheet_notebook.forget(tab_id)
        self.workbook_sheets.clear()
        self.tab_meta.clear()
        self.formula_cells.clear()
        self.df = pd.DataFrame()
        self.current_file = None
        self.current_sheet_name = None
        self.txt_sepa = ','
        self.modified = False
        self.update_title()

    def close_current_tab(self):
        """Close the active tab, asking to save if dirty."""
        self._clear_undo_for_tab(self.current_sheet_name)
        if not self.current_sheet_name:
            return
    
        meta = self.tab_meta.get(self.current_sheet_name, {})
        if meta.get("modified"):
            #fp = meta.get("file_path", "this tab")
            result = messagebox.askyesnocancel(
                "Unsaved Changes",
                f"'{self.current_sheet_name}' has unsaved changes.\n\nSave before closing?"
            )
            if result is None:    # Cancel
                return
            if result is True:    # Yes → save first
                self.save_file()
    
        # Remove from all registries
        name = self.current_sheet_name
        current_tab = self.sheet_notebook.select()
        self.sheet_notebook.forget(current_tab)
        self.workbook_sheets.pop(name, None)
        self.tab_meta.pop(name, None)
        self.formula_cells.pop(name, None)
    
        self._sv_update_combo() 
        # If no tabs left, create a blank one
        if not self.sheet_notebook.tabs():
            self.df = pd.DataFrame("",index=range(5), columns=["A", "B", "C", "D","E"])
            widget = self._create_sheet_tab("Sheet1", self.df)
            self._populate_sheet(widget, self.df)
            self.current_file = None
            self.txt_sepa = ','
            self.modified = False
            self.update_title()
        else:
            self._on_sheet_change()   # sync globals to newly active tab
    
        self.modified = any(m["modified"] for m in self.tab_meta.values())
        self.update_title()
        self.set_status(f"Closed tab: {name}")

    def new_file(self):
        """Create new empty file — replaces session after save prompt."""
        if self.modified and not self.ask_save_changes("create a new file"):
            return
        self._close_all_tabs()
        self.df = pd.DataFrame("",index=range(5), columns=["A", "B", "C", "D","E"])
        sheet_widget = self._create_sheet_tab("Sheet1", self.df)
        self._populate_sheet(sheet_widget, self.df)
        self.update_sheet_from_dataframe()
        self.update_dataframe_from_sheet()
        self.set_status("New file created")

    def add_new_sheet(self, df = None, sheet_name = None):
        """Create new empty file — replaces session after save prompt."""
        # if self.modified and not self.ask_save_changes("create a new file"):
        #     return
        # self._close_all_tabs()
        if not isinstance(df, pd.DataFrame):
            df = pd.DataFrame("",index=range(5), columns=["A", "B", "C", "D","E"])
        if not sheet_name:
            sheet_name = "Sheet1"

        sheet_widget = self._create_sheet_tab(sheet_name, df)
        self._populate_sheet(sheet_widget, df)
        self.sheet_notebook.select(len(self.sheet_notebook.tabs()) - 1)
        # Update globals to reflect current tab
        self._sync_globals_from_current_tab()
        # print("tttttttttttttt")
        # print(self.df)
        self.set_status("New file created")
        


    def update_selection_status(self, event=None):
        """Enhanced selection status with cell info."""
        sheet = self.get_current_sheet()
        if not sheet:
            return
        
        selected = sheet.get_selected_cells()
        if not selected:
            self.set_sel_status("Ready")
            self.cell_info.config(text="")
            self.formula_var.set("")
            return
        
        selected = list(selected)
        
        if len(selected) == 1:
            r, c = selected[0]
            headers = sheet.headers()
            col_name = headers[c] if c < len(headers) else f"Col{c + 1}"
            
            col_letter = self._num_to_col_letter(c)
            cell_ref = f"{col_letter}{r + 1}"
            
            cell_key = f"{r},{c}"
            sheet_formulas = self.formula_cells.get(self.current_sheet_name, {})
            if cell_key in sheet_formulas:
                self.formula_var.set(sheet_formulas[cell_key])
            else:
                value = sheet.get_cell_data(r, c)
                self.formula_var.set(value or '')
            
            self.set_sel_status(f"Cell: {cell_ref} ({col_name})")
            self.cell_info.config(text=f"Row {r + 1}, Col {c + 1}")
            return
        
        rows = [r for r, _ in selected]
        cols = [c for _, c in selected]
        row_span = max(rows) - min(rows) + 1
        col_span = max(cols) - min(cols) + 1
        
        values = []
        for r, c in selected:
            try:
                val = sheet.get_cell_data(r, c)
                if val not in ("", None):
                    values.append(float(val))
            except ValueError:
                pass
        
        stats = ""
        if values:
            count = len(values)
            s = sum(values)
            avg = s / count if count else 0
            stats = f" | Count={count}, Sum={s:.2f}, Avg={avg:.2f}"
        
        self.set_sel_status(f"Selection: {row_span}R × {col_span}C{stats}")
        self.cell_info.config(text=f"{len(selected)} cells")


    def _num_to_col_letter(self, n: int) -> str:
        """Convert column number to Excel letter."""
        result = ""
        n += 1
        while n > 0:
            n -= 1
            result = chr(65 + (n % 26)) + result
            n //= 26
        return result


    def autofill_selection(self, event=None):
        """Enhanced autofill with pattern detection."""
        sheet = self.get_current_sheet()
        if not sheet:
            return
        
        selected = list(sheet.get_selected_cells())
        if not selected:
            return
        
        rows = [r for r, _ in selected]
        cols = [c for _, c in selected]
        min_r, max_r = min(rows), max(rows)
        min_c, max_c = min(cols), max(cols)
        
        # Vertical fill
        if max_r > min_r and min_c == max_c:
            col = min_c
            values = [sheet.get_cell_data(r, col) for r in range(min_r, max_r + 1)]
            nums = []
            for v in values:
                if v and re.search(r"(\d+)$", str(v)):
                    nums.append(int(re.search(r"(\d+)$", str(v)).group(1)))
                elif v:
                    nums.append(v)
            
            step = nums[-1] - nums[-2] if len([x for x in nums if isinstance(x, int)]) >= 2 else 0
            
            last_val = None
            for i, v in enumerate(values):
                if v:
                    last_val = v
                    continue
                if isinstance(last_val, str) and re.search(r"(\d+)$", last_val):
                    prefix = re.sub(r"\d+$", "", last_val)                
                    if prefix=='-':
                        prefix = ""
                        num = int(re.search(r"(\d+)$", last_val).group(1))*-1 + step
                        new_val = f"{prefix}{num}"
                    else:
                        num = int(re.search(r"(\d+)$", last_val).group(1)) + step
                    new_val = f"{prefix}{num}"
                elif isinstance(last_val, int):
                    new_val = last_val + step
                else:
                    new_val = last_val
                values[i] = str(new_val)
                last_val = values[i]
                sheet.set_data(min_r + i, col, data=values[i], undo=True)
        
        # Horizontal fill
        elif max_c > min_c and min_r == max_r:
            row = min_r
            values = [sheet.get_cell_data(row, c) for c in range(min_c, max_c + 1)]
            nums = []
            for v in values:
                if v and re.search(r"(\d+)$", str(v)):
                    nums.append(int(re.search(r"(\d+)$", str(v)).group(1)))
                elif v:
                    nums.append(v)
            
            step = nums[-1] - nums[-2] if len([x for x in nums if isinstance(x, int)]) >= 2 else 0
            
            last_val = None
            for i, v in enumerate(values):
                if v:
                    last_val = v
                    continue
                if isinstance(last_val, str) and re.search(r"(\d+)$", last_val):
                    prefix = re.sub(r"\d+$", "", last_val)
                    if prefix=='-':
                        prefix = ""
                        num = int(re.search(r"(\d+)$", last_val).group(1))*-1 + step
                    else:
                        num = int(re.search(r"(\d+)$", last_val).group(1)) + step

                    new_val = f"{prefix}{num}"
                elif isinstance(last_val, int):
                    new_val = last_val + step
                else:
                    new_val = last_val
                values[i] = str(new_val)
                last_val = values[i]
                sheet.set_data(row, min_c + i, data=values[i], undo=True)

        self.update_dataframe_from_sheet()     
        if self.suggestion_mode_var.get():
            self.customize_turbine_columns()
        self.modified = True
        self.update_title()
        self.set_status("Autofill applied")

    def add_column(self, event=None):
        """Insert column."""
        sheet = self.get_current_sheet()
        if not sheet:
            return
        
        if event:
            col_index = sheet.get_currently_selected()[1] if sheet.get_currently_selected() else len(self.df.T)
            new_col = f"Column{len(self.df.columns) + 1}"
            # self.update_dataframe_from_sheet()
            # print('HereC1')
        else:
            col_index = len(self.df.T)
            new_col = f"Column{len(self.df.columns) + 1}"
            # self.df.insert(col_index, new_col, "")
            # self.update_sheet_from_dataframe()
            sheet.insert_column(idx=col_index, header=new_col)

        sheet.set_header_data(value=new_col, c=col_index)
        self.update_dataframe_from_sheet()
        
        self.modified = True
        self.update_title()
        self.set_status(f"Added column {new_col}")


    def insert_row(self, event=None):
        """Insert row."""
        sheet = self.get_current_sheet()
        if not sheet:
            return
        # print('Here')
        if event:
            idx = sheet.get_currently_selected()[0] if sheet.get_currently_selected() else len(self.df)
            self.update_dataframe_from_sheet()
            # print('Here1')
        else:
            idx = len(self.df)
            # empty_row = pd.Series([""] * len(self.df.columns), index=self.df.columns)
            # self.df = pd.concat([self.df.iloc[:idx], empty_row.to_frame().T, self.df.iloc[idx:]]).reset_index(drop=True)
            sheet.insert_row(idx=idx, redraw=True)
            self.update_dataframe_from_sheet()
        
        self.modified = True
        self.update_title()
        
        if self.suggestion_mode_var.get():
            self.customize_turbine_columns()
        self.set_status(f"Added row {idx+1}")
        
    def clean_whitespace(self, event=None):
        """Insert column."""
        self._push_undo()
        sheet = self.get_current_sheet()
        if not sheet:
            return
        
        if event:
            col_index = sheet.get_currently_selected()[1] if sheet.get_currently_selected() else len(self.df.T)
            col = list(self.df.columns)
            clean_col = col[col_index]
            col_name = clean_col
        else:
            col_name = None
            clean_col = 'All'
            
        self.df = trim_whitespace(df = self.df, columns = col_name)
        self.update_sheet_from_dataframe()
        
        self.modified = True
        self.mark_modified()
        self.set_status(f"Extra space cleaned on columns: {clean_col}")


    def clean_nan(self, event=None):
        """Clean NaN/None values, prompting the user for the fill value."""
        
        sheet = self.get_current_sheet()
        if not sheet:
            return
        if event:
            col_index = sheet.get_currently_selected()[1] if sheet.get_currently_selected() else len(self.df.T)
            col = list(self.df.columns)
            clean_col = col[col_index]
            col_name = [clean_col] 
        else:
            col_name = None
            clean_col = 'All'

        # The simpledialog.askstring returns None if the user cancels the dialog.
        fill_value_input = simpledialog.askstring(
            "Fill Nulls", 
            f"Enter the fill value for column: {clean_col}",
            initialvalue=''
        )

        # Check if the user cancelled the dialog
        if fill_value_input is None:
            self.set_status("NaN cleaning cancelled.")
            return

        # For now, we treat the input as a string for simplicity, as per your original " " fill_value.
        fill_val = fill_value_input if fill_value_input != "" else " " 

        # 3. Apply the cleaning function
        self._push_undo()
        self.df = fill_nulls(
            df = self.df, 
            columns = col_name, 
            fill_value = fill_val, 
            only_object_columns=False
        )
        
        self.update_sheet_from_dataframe()
        
        self.modified = True
        self.mark_modified()
        self.set_status(f"Cleaned NaN (No Undo): {clean_col} col with value '{fill_val}'")

    def change_separator(self):
        """Change CSV separator."""
        popup = tb.Toplevel(self.root)
        popup.title("Change Separator")
        popup.geometry("350x250")
        
        tb.Label(popup, text="Select separator:", font=("Segoe UI", 10, "bold")).pack(pady=10)
        
        sep_var = tk.StringVar(value=self.current_sep or self.txt_sepa)
        sep_choices = [
            ("None", "Auto"),
            (",", "Comma (,)"),
            (";", "Semicolon (;)"),
            ("\t", "Tab (\\t)"),
            (r"\s+", "Spaces(   )"),
            (" ", "Space ( )"),
            ("|", "Pipe (|)"),
            ("#label", "Library format"),
        ]
        combo = tb.Combobox(popup, values=[f"{s[0]} - {s[1]}" for s in sep_choices], textvariable=sep_var)
        combo.pack(pady=5, padx=15, fill="x")
        
        def apply_separator():
            sep_text = sep_var.get().split(" - ")[0] if " - " in sep_var.get() else sep_var.get()
            if sep_text.lower() in ["\\t", "tab"]:
                sep_text = "\t"
            elif sep_text.lower() in ["none", "auto"]:
                sep_text = None
            
            self.current_sep = sep_text
            self.txt_sepa = sep_text
            popup.destroy()
            
            if self.current_file:
                file_path = self.current_file
                self.close_current_tab()
                self.load_file(file_path = file_path, force_sep= sep_text)
            
            self.set_status(f"Separator: {repr(sep_text)}")
        
        tb.Button(popup, text="Apply", command=apply_separator, bootstyle="success").pack(pady=10)


    # MDT Features
    def toggle_suggestion_mode(self):
        """Toggle MDT suggestion mode."""
        if self.suggestion_mode_var.get():
            self.set_status("MDT Suggestion Mode: ON")
            self.customize_turbine_columns()
        else:
            self.clean_all_dropdown()
            self.set_status("MDT Suggestion Mode: OFF")


    def customize_turbine_columns(self):
        """Attach dropdown suggestions based on file type prefix."""
        if not self.current_file:
            return
        
        fname = os.path.basename(self.current_file)
        self.clean_all_dropdown()
        
        # Get dropdown config for this file type
        dropdown_config = get_mdt_dropdown_config(fname)
        
        if not dropdown_config:
            self.suggestion_mode_var.set(False)
            self.toggle_suggestion_mode()
            self.set_status("Suggestion Mode OFF (this file not supported)")
            return
        
        # Special handling for Parasolid column
        if fname.startswith("Turbine_Comp") and "Parasolid" in self.df.columns:
            folder_path = os.path.dirname(self.current_file)
            xt_files = get_file_list_for_column(folder_path, "Parasolid")
            self.set_column_dropdown("Parasolid", xt_files)
        
        # Apply all other dropdowns
        for column_name, options in dropdown_config.items():
            if column_name in self.df.columns and column_name != "Parasolid":
                self.set_column_dropdown(column_name, options)


    def set_column_dropdown(self, column_name, options):
        """Attach dropdown to all cells in a column."""
        sheet = self.get_current_sheet()
        if not sheet:
            return
        if column_name not in self.df.columns:
            return
        col_idx = self.df.columns.get_loc(column_name)
        
        if isinstance(options, dict):
            dropdown_display = [f"{val} - {desc}" for val, desc in options.items()]
            valid_values = set(options.keys())
        else:
            dropdown_display = sorted(list(options))
            valid_values = set(options)
        
        for r in range(len(self.df)):
            existing_value = sheet.get_cell_data(r, col_idx)
            sheet.dropdown(r, col_idx, values=dropdown_display, state="normal", validate_input=False)
            sheet.set_cell_data(r, col_idx, value=existing_value)
            if existing_value not in valid_values:
                sheet.highlight_cells(row=r, column=col_idx, bg="lightcoral")


    def clean_dropdown_value(self):
        """If a selected cell shows 'value - description', store only 'value'."""
        sheet = self.get_current_sheet()
        if not sheet:
            return
        selected = list(sheet.get_selected_cells())
        if len(selected) != 1:
            return
        r, c = selected[0]
        if r is None or c is None:
            return
        val = sheet.get_cell_data(r, c)
        if val and " - " in str(val):
            clean_val = str(val).split(" - ")[0].strip()
            sheet.set_cell_data(r, c, clean_val)


    def clean_all_dropdown(self):
        """Remove all dropdowns."""
        sheet = self.get_current_sheet()
        if not sheet:
            return
        
        for r in range(sheet.total_rows()):
            for c in range(sheet.total_columns()):
                sheet.del_dropdown(r, c)

    def run_component_graphs(self):
        """Plot interactive component graphs for matching Turbine_Conta_Catalogue_*.csv files."""
        if not self.current_file:
            self.set_status("No File: Please open a file first.")
            return

        root_dir = os.path.dirname(self.current_file)

        conta_files = [
            f for f in os.listdir(root_dir)
            if f.startswith("Turbine_Conta_Catalogue_") and f.endswith(".csv")
        ]
        comp_file = next(
            (f for f in os.listdir(root_dir)
             if f.startswith("Turbine_Comp_Catalogue_") and f.endswith(".csv")),
            None
        )
        comp_path = os.path.join(root_dir, comp_file) if comp_file else None

        if not conta_files:
            self.set_status("No Matching Files: No Turbine_Conta_Catalogue_*.csv files found.")
            return

        for fname in conta_files:
            fpath = os.path.join(root_dir, fname)
            out_html = os.path.join(root_dir, f"graph_{fname.split('.')[0]}.html")
            try:
                plot_interactive_component_graph(
                    fpath, output_file=out_html, open_browser=True, comp_csv=comp_path
                )
                self.set_status(f"Graph plot done: open {out_html}")
            except Exception as e:
                self.set_status(f"Graph Error: Error processing {fname}: {e}")

    def run_component_graphs_coord(self):
        """Plot interactive component graphs for matching Turbine_Conta_Catalogue_*.csv files."""
        if not self.current_file:
            self.set_status("No File: Please open a file first.")
            return

        root_dir = os.path.dirname(self.current_file)

        conta_files = [
            f for f in os.listdir(root_dir)
            if f.startswith("coordinateTable") and f.endswith(".txt")
        ]
        comp_file = next(
            (f for f in os.listdir(root_dir)
             if f.startswith("Turbine_Comp_Catalogue_") and f.endswith(".csv")),
            None
        )
        comp_path = os.path.join(root_dir, comp_file) if comp_file else None

        if not conta_files:
            self.set_status("No Matching Files: No coordinateTable*.txt files found.")
            return

        for fname in conta_files:
            fpath = os.path.join(root_dir, fname)
            out_html = os.path.join(root_dir, f"graph_{fname.split('.')[0]}.html")
            try:
                _, _, _, df = read_label_format(fpath)
                plot_interactive_component_graph_coordinate(
                    df, output_file=out_html, open_browser=True, comp_csv=comp_path
                )
                self.set_status(f"Graph plot done: open {out_html}")
            except Exception as e:
                self.set_status(f"Graph Error: Error processing {fname}: {e}")

    def show_formula_help(self):
        """Show formula help dialog."""
        help_window = tb.Toplevel(self.root)
        help_window.title("Formula Reference")
        help_window.geometry("700x600")
        
        frame = tb.Frame(help_window)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        scrollbar = tb.Scrollbar(frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        text = tk.Text(frame, wrap=tk.WORD, yscrollcommand=scrollbar.set, font=("Consolas", 10))
        text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=text.yview)
        
        help_text = FORMULA_HELP_TEXT
        
        text.insert(1.0, help_text)
        text.config(state=tk.DISABLED)
        
        tb.Button(help_window, text="Close", command=help_window.destroy, bootstyle="secondary").pack(pady=10)
        
#------------------save recent
    def _recent_path(self):
        """Full path to the recent files JSON."""
        # temp_path = Path(tempfile.gettempdir())
        pathapp = self.get_appdata_path()
        temp_file = os.path.join(pathapp, RECENT_FILE)
        os.makedirs(pathapp, exist_ok=True)
        return temp_file

    def _load_recent(self):
        """
        Return list of recent file paths (strings).
        Skips any path that no longer exists on disk.
        Returns empty list if file missing or corrupt.
        """
        try:
            with open(self._recent_path(), 'r', encoding='utf-8') as f:
                paths = json.load(f)
            # Filter out paths that have been deleted / moved
            return [p for p in paths if os.path.exists(p)]
        except Exception as e:
            return []
    
    
    def _save_recent(self, file_path):
        """
        Prepend file_path to the recent list and persist.
        Keeps at most 10 entries. Deduplicates (path already
        in the list is moved to top, not duplicated).
        """
        try:
            paths = self._load_recent()
            # Remove existing occurrence of this path if any
            paths = [p for p in paths if p != file_path]
            # Prepend the new entry
            paths.insert(0, file_path)
            # Trim to 10
            paths = paths[:10]
            print(self._recent_path())
            with open(self._recent_path(), 'w', encoding='utf-8') as f:
                json.dump(paths, f, indent=2)
        except Exception:
            print('recent file')
            pass   # recent files are non-critical — never crash for this

#------------------ col stat
    def show_column_stats(self, col_name: str = None):
        """
        Show a popup with stats for the given column.
        If col_name is None, detects from the current selection.
        """
        sheet = self.get_current_sheet()
        if not sheet:
            return
    
        # Auto-detect column from selection if not provided
        if col_name is None:
            selected = sheet.get_currently_selected()
            if not selected:
                messagebox.showinfo('Column Stats', 'Click a cell first.')
                return
            _, c = selected[0], selected[1]
            headers = sheet.headers()
            col_name = headers[c] if c < len(headers) else f'Col{c+1}'
    
        # Sync df from sheet before computing
        self.update_dataframe_from_sheet()
    
        if col_name not in self.df.columns:
            messagebox.showerror('Column Stats', f'Column "{col_name}" not found.')
            return
    
        col = self.df[col_name]
    
        # ── Compute stats ─────────────────────────────────────────────────
        total   = len(col)
        # Treat empty string as null
        nulls   = col.replace('', pd.NA).isna().sum()
        filled  = total - nulls
        unique  = col.nunique(dropna=True)
    
        # Numeric stats
        nums = pd.to_numeric(col, errors='coerce').dropna()
        is_numeric = len(nums) > 0
    
        # Top-5 most common values (for strings)
        top5 = col[col != ''].value_counts().head(5)
    
        # ── Build popup ───────────────────────────────────────────────────
        popup = tb.Toplevel(self.root)
        popup.title(f'Column Stats — {col_name}')
        popup.geometry('380x420')
        popup.resizable(True, True)
    
        # Title bar inside popup
        title_frame = tb.Frame(popup, bootstyle='primary')
        title_frame.pack(fill='x')
        tb.Label(
            title_frame,
            text=f'  {col_name}',
            font=('Segoe UI', 11, 'bold'),
            bootstyle='inverse-primary'
        ).pack(side='left', pady=6)
    
        # Stats grid
        frame = tb.Frame(popup)
        frame.pack(fill='both', expand=True, padx=14, pady=10)
    
        def row_pair(label, value, r, highlight=False):
            """Add a label-value pair in the grid."""
            bg = '#EBF3FB' if highlight else None
            lbl = tb.Label(frame, text=label, font=('Segoe UI', 9),
                           foreground='#555555', width=16, anchor='w')
            val = tb.Label(frame, text=str(value), font=('Segoe UI', 9, 'bold'),
                           anchor='w')
            lbl.grid(row=r, column=0, sticky='w', pady=2, padx=4)
            val.grid(row=r, column=1, sticky='w', pady=2, padx=4)
            if highlight:
                lbl.configure(background=bg)
                val.configure(background=bg)
    
        r = 0
        row_pair('Total rows',   total,  r); r += 1
        row_pair('Non-empty',    filled, r); r += 1
        row_pair('Empty / null', nulls,  r, highlight=(nulls>0)); r += 1
        row_pair('Unique values',unique, r); r += 1
    
        if is_numeric:
            tb.Label(frame, text='── Numeric ──', font=('Segoe UI', 9),
                     foreground='#888888').grid(
                     row=r, column=0, columnspan=2, sticky='w', pady=(10,2), padx=4)
            r += 1
            row_pair('Min',   round(float(nums.min()), 6), r); r += 1
            row_pair('Max',   round(float(nums.max()), 6), r); r += 1
            row_pair('Mean',  round(float(nums.mean()), 6), r); r += 1
            row_pair('Median',round(float(nums.median()), 6), r); r += 1
            row_pair('Std dev',round(float(nums.std()), 6) if len(nums)>1 else 'n/a', r); r += 1
    
        # Top-5 most common
        tb.Label(frame, text='── Top values ──', font=('Segoe UI', 9),
                 foreground='#888888').grid(
                 row=r, column=0, columnspan=2, sticky='w', pady=(10,2), padx=4)
        r += 1
        for val, cnt in top5.items():
            display = str(val)[:28] + '...' if len(str(val)) > 28 else str(val)
            row_pair(display, f'{cnt}×', r); r += 1
    
        tb.Button(popup, text='Close', command=popup.destroy,
                  bootstyle='secondary').pack(pady=10)
   