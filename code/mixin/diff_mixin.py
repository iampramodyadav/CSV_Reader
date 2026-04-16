"""
mixins/diff_mixin.py
====================
Data Diff — compare two open sheet tabs side-by-side.

Features
--------
* Compare any two open tabs (different files, same file different sheets,
  before/after edits — anything in workbook_sheets)
* Two alignment modes:
    - Position-based  : row 1 LEFT vs row 1 RIGHT (fast, order-dependent)
    - Key column      : align rows by a shared unique identifier column
      (correct even when rows are reordered or inserted/deleted)
* Colour coding in the result sheet:
    - White / unchanged  : identical value in both sides
    - Yellow highlight   : changed cell (LEFT = old value, RIGHT = new value)
    - Green highlight    : added row (exists only in RIGHT)
    - Red highlight      : removed row (exists only in LEFT)
* Summary bar shows: X changed  |  Y added  |  Z removed  |  N same
* Export diff result to CSV / Excel
* "Jump to next difference" button — scrolls the result sheet
* Columns that are identical across ALL rows are hidden by default
  (toggle to show them)
* The result opens as a new tab named "Diff: LEFT vs RIGHT"
  so it doesn't overwrite any existing data

Where to call it
----------------
    # In ui_components.py → Tools menu:
    tools_menu.add_command(label='Compare Two Sheets…', command=app.open_diff_dialog)

    # Optional keyboard shortcut in main_app.py:
    root.bind('<Control-d>', lambda e: app.open_diff_dialog())
"""

# import re
import pandas as pd
# import numpy as np
import tkinter as tk
from tkinter import messagebox, filedialog
import ttkbootstrap as tb
# from ttkbootstrap.constants import *
import tksheet


# ── Colour constants for diff display ─────────────────────────────────────────
CLR = {
    "changed_old": ("#FFF3CD", "#856404"),   # bg, fg  — yellow, old value
    "changed_new": ("#FFF3CD", "#856404"),   # same yellow for the pair
    "added":       ("#D4EDDA", "#155724"),   # green   — only in right
    "removed":     ("#F8D7DA", "#721C24"),   # red     — only in left
    "same":        ("#FFFFFF", "#000000"),   # white   — unchanged
    "key":         ("#E8F4FD", "#0C4A6E"),   # light blue — key column
    "header_old":  ("#FCE8B2", "#5D4037"),   # LEFT header accent
    "header_new":  ("#C8E6C9", "#1B5E20"),   # RIGHT header accent
    "header_key":  ("#B3D8F5", "#0C4A6E"),   # key column header
}


# ── Core diff engine (pure functions, no UI) ───────────────────────────────────

def _compute_diff(df_left: pd.DataFrame, df_right: pd.DataFrame,
                  key_col: str = None) -> dict:
    """
    Compute the diff between two DataFrames.

    Parameters
    ----------
    df_left, df_right : DataFrames (all values treated as strings)
    key_col           : column name to align rows by, or None for positional

    Returns
    -------
    dict with keys:
        result_df    : pd.DataFrame  — combined view ready for display
        cell_status  : dict {(row, col): status_str}
        col_headers  : list of str   — column headers for result_df
        key_col      : str | None    — key column used
        summary      : dict {changed, added, removed, same, total_cells}
        left_name    : always the original left sheet name (set by caller)
        right_name   : always the original right sheet name (set by caller)
    """
    # Normalise: all values as stripped strings, treat nan/None as ""
    def _norm(df):
        return df.astype(str).replace(
            r'^\s*$', '', regex=True
        ).replace(
            r'^nan$|^None$|^NaN$', '', regex=True
        )

    dfl = _norm(df_left.copy())
    dfr = _norm(df_right.copy())

    if key_col and key_col in dfl.columns and key_col in dfr.columns:
        
        return _key_based_diff(dfl, dfr, key_col)
    
    else:
        return _positional_diff(dfl, dfr)


def _key_based_diff(dfl, dfr, key_col):
    """Align by a shared key column then compare."""
    left_cols  = [c for c in dfl.columns]
    right_cols = [c for c in dfr.columns]
    common_data_cols = [c for c in left_cols if c in right_cols and c != key_col]
    left_only_cols   = [c for c in left_cols  if c not in right_cols and c != key_col]
    right_only_cols  = [c for c in right_cols if c not in left_cols  and c != key_col]
    print(key_col)
    # Outer merge on key column
    merged = dfl.merge(
        dfr, on=key_col, how="outer",
        suffixes=("__L", "__R"), indicator=True
    )#.fillna("")
    print(key_col)
    display_rows = []
    cell_status  = {}

    for merged_r, row in merged.iterrows():
        merge_flag = row["_merge"]   # 'both', 'left_only', 'right_only'
        row_data   = []
        display_r  = len(display_rows)
        col_offset = 0

        # ── Key column ────────────────────────────────────────────────────
        key_val = str(row[key_col]) if row[key_col] else ""
        row_data.append(key_val)
        cell_status[(display_r, col_offset)] = (
            "added" if merge_flag == "right_only"
            else "removed" if merge_flag == "left_only"
            else "key"
        )
        col_offset += 1

        # ── Common columns ────────────────────────────────────────────────
        for col in common_data_cols:
            lv = str(row.get(col + "__L", ""))
            rv = str(row.get(col + "__R", ""))

            row_data.extend([lv, rv])

            if merge_flag == "left_only":
                cell_status[(display_r, col_offset)]     = "removed"
                cell_status[(display_r, col_offset + 1)] = "removed"
            elif merge_flag == "right_only":
                cell_status[(display_r, col_offset)]     = "added"
                cell_status[(display_r, col_offset + 1)] = "added"
            elif lv != rv:
                cell_status[(display_r, col_offset)]     = "changed_old"
                cell_status[(display_r, col_offset + 1)] = "changed_new"
            else:
                cell_status[(display_r, col_offset)]     = "same"
                cell_status[(display_r, col_offset + 1)] = "same"
            col_offset += 2

        # ── Left-only columns ─────────────────────────────────────────────
        for col in left_only_cols:
            lv = str(row.get(col, ""))
            row_data.append(lv)
            cell_status[(display_r, col_offset)] = (
                "removed" if merge_flag == "left_only"
                else "same"
            )
            col_offset += 1

        # ── Right-only columns ────────────────────────────────────────────
        for col in right_only_cols:
            rv = str(row.get(col, ""))
            row_data.append(rv)
            cell_status[(display_r, col_offset)] = (
                "added" if merge_flag == "right_only"
                else "same"
            )
            col_offset += 1

        display_rows.append(row_data)

    # Build header list
    col_headers = [key_col]
    for col in common_data_cols:
        col_headers.extend([f"{col}  ←LEFT", f"{col}  RIGHT→"])
    for col in left_only_cols:
        col_headers.append(f"{col}  (LEFT only)")
    for col in right_only_cols:
        col_headers.append(f"{col}  (RIGHT only)")

    result_df = pd.DataFrame(display_rows, columns=col_headers)
    summary   = _build_summary(cell_status)
    return {
        "result_df":   result_df,
        "cell_status": cell_status,
        "col_headers": col_headers,
        "key_col":     key_col,
        "summary":     summary,
    }


def _positional_diff(dfl, dfr):
    """Align by row position (row 0 LEFT vs row 0 RIGHT)."""
    n_rows      = max(len(dfl), len(dfr))
    left_cols   = list(dfl.columns)
    right_cols  = list(dfr.columns)
    common_cols = [c for c in left_cols if c in right_cols]
    left_extra  = [c for c in left_cols  if c not in right_cols]
    right_extra = [c for c in right_cols if c not in left_cols]

    display_rows = []
    cell_status  = {}

    for r in range(n_rows):
        row_data   = []
        col_offset = 0
        display_r  = r
        has_left   = r < len(dfl)
        has_right  = r < len(dfr)

        # Common columns
        for col in common_cols:
            lv = str(dfl.iloc[r][col]) if has_left  else ""
            rv = str(dfr.iloc[r][col]) if has_right else ""
            row_data.extend([lv, rv])

            if not has_left and has_right:
                cell_status[(display_r, col_offset)]     = "added"
                cell_status[(display_r, col_offset + 1)] = "added"
            elif has_left and not has_right:
                cell_status[(display_r, col_offset)]     = "removed"
                cell_status[(display_r, col_offset + 1)] = "removed"
            elif lv != rv:
                cell_status[(display_r, col_offset)]     = "changed_old"
                cell_status[(display_r, col_offset + 1)] = "changed_new"
            else:
                cell_status[(display_r, col_offset)]     = "same"
                cell_status[(display_r, col_offset + 1)] = "same"
            col_offset += 2

        # Left-only columns
        for col in left_extra:
            lv = str(dfl.iloc[r][col]) if has_left else ""
            row_data.append(lv)
            cell_status[(display_r, col_offset)] = "removed" if not has_right else "same"
            col_offset += 1

        # Right-only columns
        for col in right_extra:
            rv = str(dfr.iloc[r][col]) if has_right else ""
            row_data.append(rv)
            cell_status[(display_r, col_offset)] = "added" if not has_left else "same"
            col_offset += 1

        display_rows.append(row_data)

    # Build headers
    col_headers = []
    for col in common_cols:
        col_headers.extend([f"{col}  ←LEFT", f"{col}  RIGHT→"])
    for col in left_extra:
        col_headers.append(f"{col}  (LEFT only)")
    for col in right_extra:
        col_headers.append(f"{col}  (RIGHT only)")

    result_df = pd.DataFrame(display_rows, columns=col_headers)
    summary   = _build_summary(cell_status)
    return {
        "result_df":   result_df,
        "cell_status": cell_status,
        "col_headers": col_headers,
        "key_col":     None,
        "summary":     summary,
    }


def _build_summary(cell_status: dict) -> dict:
    """Count diff categories from cell_status."""
    from collections import Counter
    counts = Counter(cell_status.values())
    return {
        "changed": counts.get("changed_new", 0),   # one per changed pair
        "added":   counts.get("added", 0),
        "removed": counts.get("removed", 0),
        "same":    counts.get("same", 0) + counts.get("key", 0),
        "total_cells": len(cell_status),
    }


def _find_diff_rows(cell_status: dict) -> list:
    """Return sorted list of row indices that have at least one non-same cell."""
    diff_statuses = {"changed_old", "changed_new", "added", "removed"}
    rows = sorted({r for (r, _), s in cell_status.items() if s in diff_statuses})
    return rows


# ── Mixin class ───────────────────────────────────────────────────────────────

class DiffMixin:
    """Mixin: data diff between two open sheet tabs."""

    # ── Entry point ───────────────────────────────────────────────────────────

    def open_diff_dialog(self):
        """
        Show the diff setup dialog.
        User picks LEFT tab, RIGHT tab, alignment mode, and key column.
        """
        sheet_names = list(self.workbook_sheets.keys())
        if len(sheet_names) < 2:
            messagebox.showinfo(
                "Compare Sheets",
                "You need at least two open tabs to compare.\n\n"
                "Open a second file (Ctrl+Shift+O) or add a sheet."
            )
            return

        dlg = tb.Toplevel(self.root)
        dlg.title("Compare Two Sheets")
        dlg.geometry("480x400")
        dlg.resizable(False, False)
        dlg.grab_set()

        # Header
        hdr = tb.Frame(dlg, bootstyle="primary")
        hdr.pack(fill="x")
        tb.Label(
            hdr, text="  Compare Two Sheets",
            font=("Segoe UI", 11, "bold"),
            bootstyle="inverse-primary"
        ).pack(side="left", pady=8)

        body = tb.Frame(dlg)
        body.pack(fill="both", expand=True, padx=20, pady=10)

        def lbl(text):
            tb.Label(body, text=text, font=("Segoe UI", 9, "bold")).pack(
                anchor="w", pady=(10, 2)
            )

        # LEFT sheet
        lbl("LEFT sheet  (original / baseline):")
        left_var = tk.StringVar(value=sheet_names[0])
        left_combo = tb.Combobox(
            body, textvariable=left_var,
            values=sheet_names, state="readonly", width=40
        )
        left_combo.pack(anchor="w")

        # RIGHT sheet
        lbl("RIGHT sheet  (modified / comparison):")
        right_var = tk.StringVar(
            value=sheet_names[1] if len(sheet_names) > 1 else sheet_names[0]
        )
        right_combo = tb.Combobox(
            body, textvariable=right_var,
            values=sheet_names, state="readonly", width=40
        )
        right_combo.pack(anchor="w")

        # Alignment mode
        lbl("Alignment mode:")
        mode_var = tk.StringVar(value="position")
        mode_frame = tb.Frame(body)
        mode_frame.pack(anchor="w")
        tb.Radiobutton(
            mode_frame, text="By row position  (fast, rows must be in same order)",
            variable=mode_var, value="position"
        ).pack(anchor="w")
        tb.Radiobutton(
            mode_frame,
            text="By key column  (handles added / removed / reordered rows)",
            variable=mode_var, value="key"
        ).pack(anchor="w")

        # Key column
        key_frame = tb.Frame(body)
        key_frame.pack(anchor="w", pady=(4, 0))
        tb.Label(key_frame, text="Key column:", font=("Segoe UI", 9)).pack(
            side="left"
        )
        key_var = tk.StringVar()
        key_combo = tb.Combobox(
            key_frame, textvariable=key_var, width=24, state="readonly"
        )
        key_combo.pack(side="left", padx=8)

        def update_key_choices(*_):
            # Offer columns shared between both selected sheets
            ln = left_var.get()
            rn = right_var.get()
            dfl = self.workbook_sheets.get(ln, pd.DataFrame())
            dfr = self.workbook_sheets.get(rn, pd.DataFrame())
            shared = [c for c in dfl.columns if c in dfr.columns]
            key_combo["values"] = shared
            if shared:
                key_var.set(shared[0])

        left_var.trace_add("write", update_key_choices)
        right_var.trace_add("write", update_key_choices)
        mode_var.trace_add("write", lambda *_: key_combo.configure(
            state="readonly" if mode_var.get() == "key" else "disabled"
        ))
        update_key_choices()

        # Buttons
        btn_row = tb.Frame(body)
        btn_row.pack(pady=16)

        def run_diff():
            ln = left_var.get()
            rn = right_var.get()
            if ln == rn:
                messagebox.showwarning(
                    "Compare Sheets",
                    "LEFT and RIGHT must be different tabs."
                )
                return
            key = key_var.get() if mode_var.get() == "key" else None
            dlg.destroy()
            self._run_diff(ln, rn, key_col=key)

        tb.Button(btn_row, text="Compare →",  command=run_diff,
                  bootstyle="success", width=14).pack(side="left", padx=6)
        tb.Button(btn_row, text="Cancel",     command=dlg.destroy,
                  bootstyle="secondary", width=10).pack(side="left", padx=6)

        dlg.bind("<Return>", lambda e: run_diff())

    # ── Main diff runner ──────────────────────────────────────────────────────

    def _run_diff(self, left_name: str, right_name: str, key_col: str = None):
        """
        Compute the diff and open a new tab showing the result.
        Called after the setup dialog is confirmed.
        """
        # Sync current sheet data first
        self.update_dataframe_from_sheet()

        dfl = self.workbook_sheets.get(left_name,  pd.DataFrame())
        dfr = self.workbook_sheets.get(right_name, pd.DataFrame())

        if dfl.empty and dfr.empty:
            messagebox.showinfo("Compare Sheets", "Both sheets are empty.")
            return

        self.set_status("Computing diff…")
        self.root.update_idletasks()

        try:
            diff = _compute_diff(dfl, dfr, key_col=key_col)
        except Exception as exc:
            messagebox.showerror("Compare Sheets", f"Diff failed:\n{exc}")
            self.set_status("Diff error")
            return

        diff["left_name"]  = left_name
        diff["right_name"] = right_name

        # Build the diff result window (separate Toplevel, not a tab)
        self._open_diff_window(diff)

        s = diff["summary"]
        self.set_status(
            f"Diff done — {s['changed']} changed  |  "
            f"{s['added']} added  |  {s['removed']} removed  |  "
            f"{s['same']} same"
        )

    # ── Result window ─────────────────────────────────────────────────────────

    def _open_diff_window(self, diff: dict):
        """
        Open a new Toplevel window showing the diff result.
        Uses tksheet for the data grid so it feels identical to the main app.
        """
        result_df   = diff["result_df"]
        cell_status = diff["cell_status"]
        col_headers = diff["col_headers"]
        summary     = diff["summary"]
        left_name   = diff["left_name"]
        right_name  = diff["right_name"]
        key_col     = diff["key_col"]

        win = tb.Toplevel(self.root)
        win.title(f"Diff:  {left_name}  ←→  {right_name}")
        win.geometry("1100x680")
        win.resizable(True, True)

        # ── Top bar ───────────────────────────────────────────────────────
        top = tb.Frame(win, bootstyle="dark")
        top.pack(fill="x")

        tb.Label(
            top,
            text=f"  {left_name}  ←→  {right_name}"
            + (f"  [key: {key_col}]" if key_col else "  [positional]"),
            font=("Segoe UI", 10, "bold"),
            bootstyle="inverse-dark"
        ).pack(side="left", pady=6, padx=4)

        # ── Summary bar ───────────────────────────────────────────────────
        summary_frame = tb.Frame(win)
        summary_frame.pack(fill="x", padx=8, pady=(6, 2))

        def summary_badge(text, bg, fg):
            lbl = tk.Label(
                summary_frame, text=f"  {text}  ",
                bg=bg, fg=fg,
                font=("Segoe UI", 9, "bold"),
                relief="flat", padx=6, pady=3
            )
            lbl.pack(side="left", padx=4)

        summary_badge(f"✎ {summary['changed']} changed",  "#FFF3CD", "#856404")
        summary_badge(f"+ {summary['added']} added",       "#D4EDDA", "#155724")
        summary_badge(f"− {summary['removed']} removed",   "#F8D7DA", "#721C24")
        summary_badge(f"= {summary['same']} same",         "#F0F0F0", "#444444")

        total_diffs = summary["changed"] + summary["added"] + summary["removed"]
        if total_diffs == 0:
            summary_badge("✓ Sheets are identical", "#D4EDDA", "#155724")

        # ── Toolbar ───────────────────────────────────────────────────────
        toolbar = tb.Frame(win)
        toolbar.pack(fill="x", padx=8, pady=2)

        # Navigation state
        diff_rows = _find_diff_rows(cell_status)
        nav_state = {"idx": -1}

        def _jump_next():
            if not diff_rows:
                return
            nav_state["idx"] = (nav_state["idx"] + 1) % len(diff_rows)
            sheet.see(diff_rows[nav_state["idx"]], 0)
            sheet.select_row(diff_rows[nav_state["idx"]])
            nav_lbl.config(
                text=f"Diff {nav_state['idx'] + 1} / {len(diff_rows)}"
            )

        def _jump_prev():
            if not diff_rows:
                return
            nav_state["idx"] = (nav_state["idx"] - 1) % len(diff_rows)
            sheet.see(diff_rows[nav_state["idx"]], 0)
            sheet.select_row(diff_rows[nav_state["idx"]])
            nav_lbl.config(
                text=f"Diff {nav_state['idx'] + 1} / {len(diff_rows)}"
            )

        # Hide same columns toggle
        hide_same_var = tk.BooleanVar(value=True)
        same_cols = _find_same_columns(cell_status, len(result_df), len(col_headers))

        tb.Button(toolbar, text="◀ Prev diff", command=_jump_prev,
                  bootstyle="outline-secondary", width=12).pack(side="left", padx=3)
        tb.Button(toolbar, text="Next diff ▶", command=_jump_next,
                  bootstyle="outline-secondary", width=12).pack(side="left", padx=3)

        nav_lbl = tb.Label(
            toolbar,
            text=f"{len(diff_rows)} diff row(s)" if diff_rows else "No differences",
            font=("Segoe UI", 9)
        )
        nav_lbl.pack(side="left", padx=10)

        tb.Separator(toolbar, orient="vertical").pack(
            side="left", fill="y", padx=8, pady=4
        )

        tb.Checkbutton(
            toolbar,
            text="Hide identical columns",
            variable=hide_same_var,
            command=lambda: _toggle_same_cols(),
            bootstyle="round-toggle"
        ).pack(side="left", padx=4)

        tb.Separator(toolbar, orient="vertical").pack(
            side="left", fill="y", padx=8, pady=4
        )

        tb.Button(toolbar, text="Export CSV",   command=lambda: _export("csv"),
                  bootstyle="outline-primary", width=12).pack(side="right", padx=3)
        tb.Button(toolbar, text="Export Excel", command=lambda: _export("xlsx"),
                  bootstyle="outline-success", width=12).pack(side="right", padx=3)

        # ── Legend ────────────────────────────────────────────────────────
        legend_frame = tb.Frame(win)
        legend_frame.pack(fill="x", padx=8, pady=(0, 4))

        def legend_swatch(bg, fg, text):
            tk.Label(
                legend_frame, text=f" {text} ",
                bg=bg, fg=fg, font=("Segoe UI", 8),
                relief="solid", bd=1, padx=4, pady=1
            ).pack(side="left", padx=4)

        legend_swatch(CLR["changed_old"][0], CLR["changed_old"][1], "Changed")
        legend_swatch(CLR["added"][0],       CLR["added"][1],       "Added (only in RIGHT)")
        legend_swatch(CLR["removed"][0],     CLR["removed"][1],     "Removed (only in LEFT)")
        legend_swatch(CLR["same"][0],        CLR["same"][1],        "Unchanged")
        legend_swatch(CLR["key"][0],         CLR["key"][1],         "Key column")

        # ── Sheet widget ──────────────────────────────────────────────────
        sheet_frame = tb.Frame(win)
        sheet_frame.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        sheet = tksheet.Sheet(
            sheet_frame,
            show_x_scrollbar=True, show_y_scrollbar=True,
            show_top_left=True, headers=None
        )
        sheet.enable_bindings((
            "single_select", "row_select", "column_select", "drag_select",
            "copy", "arrowkeys", "find", "column_width_resize",
            "row_height_resize"
        ))
        sheet.pack(fill="both", expand=True)

        # Populate data
        data = result_df.astype(str).values.tolist()
        sheet.headers(col_headers)
        sheet.set_sheet_data(data)
        sheet.set_options(align="center")

        # Apply colour highlights
        _apply_highlights(sheet, cell_status, key_col, col_headers)

        # Column widths — slightly wider for diff view
        for c in range(len(col_headers)):
            sheet.column_width(column=c, width=130)

        # Apply initial "hide same columns" state
        _hidden_cols = {"cols": []}

        def _toggle_same_cols():
            if hide_same_var.get():
                # Hide columns that are entirely "same"
                hide_indices = same_cols
                _hidden_cols["cols"] = hide_indices
                for c in range(len(col_headers)):
                    if c in hide_indices:
                        sheet.column_width(column=c, width=0)
                    else:
                        sheet.column_width(column=c, width=130)
            else:
                _hidden_cols["cols"] = []
                for c in range(len(col_headers)):
                    sheet.column_width(column=c, width=130)
            sheet.refresh()

        _toggle_same_cols()   # apply on open

        # ── Export helper ─────────────────────────────────────────────────
        def _export(fmt):
            fp = filedialog.asksaveasfilename(
                title="Export Diff Result",
                defaultextension=f".{fmt}",
                filetypes=[
                    ("CSV", "*.csv") if fmt == "csv"
                    else ("Excel", "*.xlsx")
                ],
                initialfile=f"diff_{left_name}_vs_{right_name}.{fmt}"
            )
            if not fp:
                return
            try:
                if fmt == "csv":
                    result_df.to_csv(fp, index=False)
                else:
                    with pd.ExcelWriter(fp, engine="openpyxl") as writer:
                        result_df.to_excel(writer, sheet_name="Diff", index=False)
                        # Apply Excel cell colours
                        _apply_excel_colours(
                            writer, "Diff", cell_status,
                            len(result_df), len(col_headers)
                        )
                messagebox.showinfo("Export", f"Saved:\n{fp}")
            except Exception as exc:
                messagebox.showerror("Export Error", str(exc))


# ── Highlight helpers ─────────────────────────────────────────────────────────

def _apply_highlights(sheet, cell_status: dict, key_col, col_headers):
    """Apply tksheet highlight_cells for every cell in the diff result."""
    for (r, c), status in cell_status.items():
        if status in CLR:
            bg, fg = CLR[status]
            sheet.highlight_cells(row=r, column=c, bg=bg, fg=fg)
    sheet.refresh()


def _find_same_columns(cell_status: dict, n_rows: int, n_cols: int) -> list:
    """
    Return column indices where EVERY cell is 'same' or 'key'.
    These are candidates for hiding in the diff view.
    """
    diff_statuses = {"changed_old", "changed_new", "added", "removed"}
    diff_cols = {c for (_, c), s in cell_status.items() if s in diff_statuses}
    return [c for c in range(n_cols) if c not in diff_cols]


def _apply_excel_colours(writer, sheet_name: str, cell_status: dict,
                         n_rows: int, n_cols: int):
    """Apply openpyxl fill colours to the exported Excel diff."""
    try:
        from openpyxl.styles import PatternFill, Font

        # Hex → openpyxl fill map
        STATUS_FILL = {
            "changed_old": ("FFF3CD", "856404"),
            "changed_new": ("FFF3CD", "856404"),
            "added":       ("D4EDDA", "155724"),
            "removed":     ("F8D7DA", "721C24"),
            "key":         ("E8F4FD", "0C4A6E"),
        }

        ws = writer.sheets[sheet_name]

        for (r, c), status in cell_status.items():
            if status in STATUS_FILL:
                bg_hex, fg_hex = STATUS_FILL[status]
                fill = PatternFill(
                    start_color=bg_hex, end_color=bg_hex, fill_type="solid"
                )
                font = Font(color=fg_hex)
                # openpyxl is 1-indexed; row 0 = row 2 (row 1 = header)
                cell = ws.cell(row=r + 2, column=c + 1)
                cell.fill = fill
                cell.font = font

    except ImportError:
        pass   # openpyxl not available — colours skipped, data still exports